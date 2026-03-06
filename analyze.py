import os
import requests
import json
import ssl
from datetime import datetime, timedelta
import time
import traceback
import pandas as pd
import concurrent.futures
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import math
import argparse
import sys
import io

# Force UTF-8 encoding for stdout/stderr to fix Chinese character corruption in Windows console
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7',
    'Connection': 'keep-alive',
    'Referer': 'https://www.tpex.org.tw/'
}

def get_json(url):
    retries = 3
    # Create a local headers dictionary to ensure Referer is always present for TPEX calls
    # and to avoid modifying the global headers for specific calls if needed later.
    # For TWSE, the global headers are fine.
    local_headers = headers.copy()
    if "tpex.org.tw" in url:
        local_headers['Referer'] = 'https://www.tpex.org.tw/'
    elif "twse.com.tw" in url:
        local_headers['Referer'] = 'https://www.twse.com.tw/'

    for attempt in range(retries):
        try:
            session = requests.Session()
            # Disable SSL verification to prevent "CERTIFICATE VERIFY FAILED" on some Linux/Docker environments like Render
            # Increased timeout to 30s and added retry logic to handle transient TWSE network lag
            res = session.get(url, headers=local_headers, timeout=30, verify=False)
            # Check HTTP response status and throw if not 200
            res.raise_for_status()
            return res.json()
        except Exception as e:
            if attempt < retries - 1:
                wait_time = (attempt + 1) * 2
                print(f"Attempt {attempt + 1} failed for {url}: {e}. Retrying in {wait_time}s...")
                time.sleep(wait_time)
            else:
                print(f"Error fetching {url} after {retries} attempts: {e}")
                return None

def validate_trading_day(date_str):
    """
    使用體積極小的 '市場成交概況' API 來快速預檢當天是否為有效交易日。
    這比直接抓整份法人買賣超 (T86) 輕量得多，適合用來做前置測試。
    """
    # MI_INDEX type=MS 是市場成交概況，回傳資料極少
    url = f"https://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&date={date_str}&type=MS"
    data = get_json(url)
    # 如果 data['stat'] 為 'OK'，代表當天有交易紀錄
    return data and data.get('stat') == 'OK'

def fetch_twse(date="20260223"):
    t86_url = f"https://www.twse.com.tw/fund/T86?response=json&date={date}&selectType=ALL"
    mi_url = f"https://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&date={date}&type=ALLBUT0999"
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        f_t86 = executor.submit(get_json, t86_url)
        f_mi = executor.submit(get_json, mi_url)
        t86_data = f_t86.result()
        mi_data = f_mi.result()
    
    if not t86_data or 'data' not in t86_data:
        print("Failed to get TWSE T86")
        return []
        
    prices = {}
    if mi_data and 'tables' in mi_data:
        # MI_INDEX tables structure, usually the 9th table is closing prices
        target_table = None
        for table in mi_data['tables']:
            fields = table.get('fields', [])
            if '證券代號' in fields and '收盤價' in fields:
                target_table = table
                break
        
        if target_table and 'fields' in target_table:
            mi_fields = target_table['fields']
            try:
                idx_mi_code = mi_fields.index('證券代號')
                idx_mi_close = mi_fields.index('收盤價')
                idx_mi_vol = mi_fields.index('成交股數')
                idx_mi_val = mi_fields.index('成交金額')
                # Add indices for change calculation
                idx_mi_sign = mi_fields.index('漲跌(+/-)')
                idx_mi_diff = mi_fields.index('漲跌價差')
                
                for row in target_table['data']:
                    code = row[idx_mi_code].strip()
                    price_str = row[idx_mi_close].replace(',', '')
                    vol_str = row[idx_mi_vol].replace(',', '')
                    val_str = row[idx_mi_val].replace(',', '')
                    sign_html = row[idx_mi_sign] # contains <p style="color:red">+</p> or similar
                    diff_str = row[idx_mi_diff].replace(',', '')
                    
                    try:
                        close_p = float(price_str)
                    except ValueError:
                        close_p = 0.0
                        
                    # Calculate change percentage
                    change_pct = 0.0
                    try:
                        diff = float(diff_str)
                        # Extract sign from HTML-like string
                        if '+' in sign_html or 'red' in sign_html:
                            pass # diff is positive
                        elif '-' in sign_html or 'green' in sign_html:
                            diff = -diff
                        
                        prev_close = close_p - diff
                        if prev_close > 0:
                            change_pct = (diff / prev_close) * 100
                    except:
                        pass

                    vwap = close_p
                    if vol_str.isdigit() and val_str.isdigit():
                        vol = int(vol_str)
                        val = int(val_str)
                        if vol > 0:
                            vwap = val / vol
                    
                    prices[code] = {
                        'close': close_p, 
                        'vwap': round(vwap, 1), 
                        'change_pct': round(change_pct, 2)
                    }
            except Exception as e:
                print("Error parsing TWSE MI_INDEX fields:", e)
    
    results = []
    t86_fields = t86_data['fields']
    idx_code = t86_fields.index('證券代號')
    idx_name = t86_fields.index('證券名稱')
    
    try:
        idx_foreign = next(i for i, f in enumerate(t86_fields) if '外陸資買賣超股數(不含外資自營商)' in f)
    except StopIteration:
        idx_foreign = next(i for i, f in enumerate(t86_fields) if '外資' in f and '買賣超' in f)
        
    try:
        idx_it = next(i for i, f in enumerate(t86_fields) if f == '投信買賣超股數')
    except StopIteration:
        idx_it = next(i for i, f in enumerate(t86_fields) if '投信' in f and '買賣超' in f)

    import math
    for row in t86_data['data']:
        code = row[idx_code].strip()
        name = row[idx_name].strip()
        
        # If no price or price is 0, we can calculate value
        if code not in prices or prices[code]['close'] == 0:
            continue
            
        # 排除 ETF 與非普通股 (代號長度非 4 或 0 開頭)
        if len(code) != 4 or code.startswith('0'):
            continue
            
        try:
            foreign_shares = int(row[idx_foreign].replace(',', ''))
            it_shares = int(row[idx_it].replace(',', ''))
        except ValueError:
            continue
            
        foreign_value = foreign_shares * prices[code]['vwap']
        it_value = it_shares * prices[code]['vwap']
        
        results.append({
            'market': 'TWSE',
            'code': code,
            'name': name,
            'price': prices[code]['close'],
            'change_pct': prices[code]['change_pct'],
            'vwap': prices[code]['vwap'],
            'foreign_val': foreign_value,
            'it_val': it_value,
            'foreign_shares': foreign_shares,
            'it_shares': it_shares,
            'foreign_lots': math.ceil(abs(foreign_shares) / 1000.0) * (1 if foreign_shares >= 0 else -1),
            'it_lots': math.ceil(abs(it_shares) / 1000.0) * (1 if it_shares >= 0 else -1)
        })
        
    return results

def fetch_tpex(date_roc="115/02/23"):
    # tpex T86 equivalent
    t86_url = f"https://www.tpex.org.tw/web/stock/3insti/daily_trade/3itrade_hedge_result.php?l=zh-tw&se=EW&t=D&d={date_roc}"
    # tpex MI_INDEX equivalent - using stk_wn1430_result.php with se=AL to get all stocks for history
    mi_url = f"https://www.tpex.org.tw/web/stock/aftertrading/otc_quotes_no1430/stk_wn1430_result.php?l=zh-tw&d={date_roc}&se=AL"
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        f_t86 = executor.submit(get_json, t86_url)
        f_mi = executor.submit(get_json, mi_url)
        t86_data = f_t86.result()
        mi_data = f_mi.result()
    
    results = []
    if not t86_data or 'tables' not in t86_data or not t86_data['tables']:
        print("Failed to get TPEX T86")
        return results
        
    if not mi_data or 'tables' not in mi_data or not mi_data['tables']:
        print("Failed to get TPEX closing prices")
        return results
        
    prices = {}
    # stk_wn1430_result.php uses 'aaData' key for the list
    mi_list = mi_data.get('aaData', [])
    if not mi_list and 'tables' in mi_data and mi_data['tables']:
        mi_list = mi_data['tables'][0].get('data', [])
        
    for row in mi_list:
        code = row[0].strip()
        price_str = str(row[2]).replace(',', '') # idx 2 is '收盤'
        diff_str = str(row[3]).replace(',', '').replace(' ', '') # idx 3 is '漲跌'
        
        try:
            close_p = float(price_str)
        except ValueError:
            close_p = 0.0
            
        # Calculate change percentage
        change_pct = 0.0
        try:
            diff = float(diff_str)
            prev_close = close_p - diff
            if prev_close > 0:
                change_pct = (diff / prev_close) * 100
        except:
            pass
            
        vwap = close_p
        try:
            # TPEX MI_INDEX: idx 8 is volume (股), idx 9 is value (元), idx 7 is vwap (if exists)
                vol_str = str(row[7]).replace(',', '')
                val_str = str(row[8]).replace(',', '')
                if vol_str.replace('.', '').isdigit() and val_str.replace('.', '').isdigit() and float(vol_str) > 0:
                    vwap = float(val_str) / float(vol_str)
                elif str(row[7]).replace('.', '').replace(',', '').isdigit():
                    vwap = float(str(row[7]).replace(',', ''))
        except Exception:
            pass
            
        prices[code] = {
            'close': close_p, 
            'vwap': round(vwap, 1), 
            'change_pct': round(change_pct, 2)
        }
            
    t86_table = t86_data['tables'][0]
    import math
    for row in t86_table.get('data', []):
        # TPEX format: 0=代號, 1=名稱
        # 4=外資買賣超, 7=外資自營買賣超, 10=外資合計, 13=投信買賣超
        code = str(row[0]).strip()
        name = str(row[1]).strip()
        if code not in prices or prices[code]['close'] == 0:
            continue
            
        # 排除 ETF 與非普通股 (代號長度非 4 或 0 開頭)
        if len(code) != 4 or code.startswith('0'):
            continue
            
        try:
            foreign_shares = int(str(row[4]).replace(',', ''))
            it_shares = int(str(row[13]).replace(',', ''))
        except (ValueError, IndexError):
            continue
            
        foreign_value = foreign_shares * prices[code]['vwap']
        it_value = it_shares * prices[code]['vwap']
        
        results.append({
            'market': 'TPEX',
            'code': code,
            'name': name,
            'price': prices[code]['close'],
            'change_pct': prices[code]['change_pct'],
            'vwap': prices[code]['vwap'],
            'foreign_val': foreign_value,
            'it_val': it_value,
            'foreign_shares': foreign_shares,
            'it_shares': it_shares,
            'foreign_lots': math.ceil(abs(foreign_shares) / 1000.0) * (1 if foreign_shares >= 0 else -1),
            'it_lots': math.ceil(abs(it_shares) / 1000.0) * (1 if it_shares >= 0 else -1)
        })
    
    return results

def format_val(val):
    if val >= 0:
        return f"+{val/100000000:.2f}億元"
    else:
        return f"{val/100000000:.2f}億元"

def analyze(target_date_str=None, period='day'):
    if not target_date_str:
        target_date_str = datetime.now().strftime('%Y%m%d')
        
    year = int(target_date_str[:4])
    month = target_date_str[4:6]
    day = target_date_str[6:8]
    roc_year = year - 1911
    
    twse_date = target_date_str
    
    dates_to_process = [target_date_str]
    
    if period == 'week':
        # Find all trading days in the same week as target_date_str
        target_dt = datetime.strptime(target_date_str, '%Y%m%d')
        # Get Monday of that week
        monday = target_dt - timedelta(days=target_dt.weekday())
        dates_to_process = []
        for i in range(5): # Mon-Fri
            d_str = (monday + timedelta(days=i)).strftime('%Y%m%d')
            # Check if it was a trading day (we skip weekends and holidays)
            if validate_trading_day(d_str):
                dates_to_process.append(d_str)
        
        print(f"Weekly analysis triggered. Processing {len(dates_to_process)} days: {dates_to_process}")

    # Gather data for all targeted dates
    all_daily_data = [] # List of lists of stock dicts
    
    for d_str in dates_to_process:
        print(f"--- Fetching data for {d_str} ---")
        # Convert YYYYMMDD to ROC date for TPEX
        current_year = int(d_str[:4])
        current_month = d_str[4:6]
        current_day = d_str[6:8]
        roc_year = current_year - 1911
        tpex_date_roc = f"{roc_year:03d}/{current_month}/{current_day}"

        twse_data = fetch_twse(d_str)
        tpex_data = fetch_tpex(tpex_date_roc)
        all_daily_data.append(twse_data + tpex_data)
        
    if not all_daily_data or not any(all_daily_data):
        print(f"No data found for the specified period.")
        return

    # Aggregation
    aggregated = {} # code -> aggregated_data
    latest_prices = {} # code -> price_info
    
    for daily_list in all_daily_data:
        for stock in daily_list:
            code = stock['code']
            if code not in aggregated:
                aggregated[code] = {
                    'market': stock['market'],
                    'code': code,
                    'name': stock['name'],
                    'foreign_shares': 0,
                    'it_shares': 0,
                    'foreign_val': 0.0,
                    'it_val': 0.0,
                    'daily_vwaps': [], # To store daily VWAPs for weekly average
                    'baseline_price': None # Weekly baseline (prev week close)
                }
            
            agg = aggregated[code]
            
            # For weekly analysis, we need the closing price of the day BEFORE this week started
            # We can derive it from the FIRST day's close and its daily change
            if period == 'week' and agg['baseline_price'] is None:
                try:
                    # Baseline = Current_Close - Daily_Change
                    # Note: stock['change_pct'] is (change / prev_close) * 100
                    # So prev_close = Current_Close / (1 + change_pct/100)
                    agg['baseline_price'] = stock['price'] / (1 + stock['change_pct'] / 100.0)
                except ZeroDivisionError:
                    agg['baseline_price'] = stock['price']
            
            agg['foreign_shares'] += stock['foreign_shares']
            agg['it_shares'] += stock['it_shares']
            agg['foreign_val'] += stock['foreign_val']
            agg['it_val'] += stock['it_val']
            agg['daily_vwaps'].append(stock['vwap'])
            
            # For VWAP we need absolute totals or based on shares?
            # Actually our 'vwap' from fetch is per share. 
            # We don't have total daily market volume/value easily without fetching more.
            # But we can approximate Weekly VWAP as: Σ(Daily_VWAP * Abs(Inst_Shares)) / Σ(Abs(Inst_Shares))
            # Or simpler: just use the sum of values and shares if we think inst_shares is the base.
            # But the 'foreign_val' we calculated is already foreign_shares * vwap.
            # So Weekly_Foreign_VWAP = Σ(Daily_Foreign_Val) / Σ(Daily_Foreign_Shares)
            
            # Let's keep a record of latest price and change calculation
            latest_prices[code] = {
                'price': stock['price'],
                'change_pct': stock['change_pct'], # This will be overwritten by later days
                'vwap': stock['vwap'] # This will be overwritten by later days
            }

    results = []
    for code, agg in aggregated.items():
        # Final day closing price and cumulative weekly change percentage
        # Correct Weekly Change% = (Friday_Close - PrevFriday_Close) / PrevFriday_Close
        # Since we don't fetch PrevFriday, we can either use the sum of change_pct (rough) 
        # or just the latest day's daily change. 
        # For simplicity and accuracy of "current state", we use the latest day info.
        
        info = latest_prices[code]
        
        # Recalculate VWAP for the whole week if period is week
        # Use simple mean of VWAPs for now, or weighted if we had total volume.
        # Given our data: Weekly_VWAP = Σ(Daily_VWAP) / Count
        # Or better: if we have total insta value and shares:
        f_shares = agg['foreign_shares']
        i_shares = agg['it_shares']
        
        final_vwap = info['vwap']
        if period == 'week' and agg['daily_vwaps']:
            # Simple average of daily VWAPs
            final_vwap = round(sum(agg['daily_vwaps']) / len(agg['daily_vwaps']), 1)

        # Weekly Change Calculation
        if period == 'week' and agg.get('baseline_price'):
            final_change_pct = round(((info['price'] - agg['baseline_price']) / agg['baseline_price']) * 100, 2)
        else:
            final_change_pct = info['change_pct']

        results.append({
            'market': agg['market'],
            'code': code,
            'name': agg['name'],
            'price': info['price'],
            'change_pct': final_change_pct,
            'vwap': final_vwap,
            'foreign_val': agg['foreign_val'],
            'it_val': agg['it_val'],
            'total_inst_val': agg['foreign_val'] + agg['it_val'],
            'foreign_shares': f_shares,
            'it_shares': i_shares,
            'foreign_lots': math.ceil(abs(f_shares) / 1000.0) * (1 if f_shares >= 0 else -1),
            'it_lots': math.ceil(abs(i_shares) / 1000.0) * (1 if i_shares >= 0 else -1)
        })
    
    all_data = results
    # Continue with ranking and report generation...
    # Let's keep it simple: if all_data exists, it returns True at the end of function
    # Wait, I see analyze function doesn't return anything. I'll modify it to return success status.

        
    print(f"Successfully processed {len(all_data)} stocks.")
    print("="*60)
    
    # Sort by foreign value
    foreign_buy = sorted([d for d in all_data if d['foreign_val'] > 0], key=lambda x: x['foreign_val'], reverse=True)
    foreign_sell = sorted([d for d in all_data if d['foreign_val'] < 0], key=lambda x: x['foreign_val'])
    
    # Sort by IT value
    it_buy = sorted([d for d in all_data if d['it_val'] > 0], key=lambda x: x['it_val'], reverse=True)
    it_sell = sorted([d for d in all_data if d['it_val'] < 0], key=lambda x: x['it_val'])
    
    print("\n### 外資買超排名 (依成交值)")
    for i, d in enumerate(foreign_buy[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : {format_val(d['foreign_val'])}")
        
    print("\n### 外資賣超排名 (依成交值)")
    for i, d in enumerate(foreign_sell[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : {format_val(d['foreign_val'])}")
        
    print("\n### 投信買超排名 (依成交值)")
    for i, d in enumerate(it_buy[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : {format_val(d['it_val'])}")
        
    print("\n### 投信賣超排名 (依成交值)")
    for i, d in enumerate(it_sell[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : {format_val(d['it_val'])}")

    print("\n" + "="*60)
    
    # 同向與反向分析
    # 同向買超: 外資買超 > 0 且 投信買超 > 0, 依加總值排序
    same_buy = [d for d in all_data if d['foreign_val'] > 0 and d['it_val'] > 0]
    same_buy.sort(key=lambda x: x['foreign_val'] + x['it_val'], reverse=True)
    
    print("\n### 土洋同買超 (外資與投信皆買超，依總買超金額排序)")
    for i, d in enumerate(same_buy[:10], 1):
        total = d['foreign_val'] + d['it_val']
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : 總計 {format_val(total)} (外資 {format_val(d['foreign_val'])}, 投信 {format_val(d['it_val'])})")
        
    # 同向賣超: 外資賣超 < 0 且 投信賣超 < 0
    same_sell = [d for d in all_data if d['foreign_val'] < 0 and d['it_val'] < 0]
    same_sell.sort(key=lambda x: x['foreign_val'] + x['it_val'])
    
    print("\n### 土洋同賣超 (外資與投信皆賣超，依總賣超金額排序)")
    for i, d in enumerate(same_sell[:10], 1):
        total = d['foreign_val'] + d['it_val']
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : 總計 {format_val(total)} (外資 {format_val(d['foreign_val'])}, 投信 {format_val(d['it_val'])})")
        
    print("\n" + "="*60)
    
    # 土洋對作: 外資與投信方向相反
    # 分為: 外資買/投信賣, 外資賣/投信買 (依兩者絕對值加總排序表示激烈程度)
    opp_fb_is = [d for d in all_data if d['foreign_val'] > 0 > d['it_val']]
    opp_fb_is.sort(key=lambda x: abs(x['foreign_val']) + abs(x['it_val']), reverse=True)
    
    print("\n### 土洋對作: 外資買超、投信賣超 (依對作規模排序)")
    for i, d in enumerate(opp_fb_is[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : 外資 {format_val(d['foreign_val'])}, 投信 {format_val(d['it_val'])}")
        
    opp_fs_ib = [d for d in all_data if d['foreign_val'] < 0 < d['it_val']]
    opp_fs_ib.sort(key=lambda x: abs(x['foreign_val']) + abs(x['it_val']), reverse=True)
    
    print("\n### 土洋對作: 外資賣超、投信買超 (依對作規模排序)")
    for i, d in enumerate(opp_fs_ib[:10], 1):
        print(f"{i:2d}. {d['code']:<6} {d['name']:<10} : 外資 {format_val(d['foreign_val'])}, 投信 {format_val(d['it_val'])}")
        
    print("\n" + "="*60 + "\n完成！")

    # 輸出成四欄位、分上市櫃的 Excel 報表
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        import subprocess
    # openpyxl is already imported at the top
    try:
        wb = Workbook()
        if len(wb.sheetnames) > 0:
            for sheet in wb.sheetnames:
                wb.remove(wb[sheet])
        
        # 建立格式與字體
        # report_date = f"{year}/{month}/{day}" # Original line, now handled by header_text
        
        if period == 'week':
            filename = f"market_analysis_week_{target_date_str}.xlsx"
            sheet_title = f"{target_date_str} 週報 (大戶動態)"
            header_text = f"台股三大法人週報分析 - 基準日期: {target_date_str}" 
            report_date = f"週報: {target_date_str}"
        else:
            filename = f"market_analysis_{target_date_str}.xlsx"
            sheet_title = f"{target_date_str} 分析 (大戶動態)"
            header_text = f"台股三大法人大戶動向分析 - 日期: {target_date_str}"
            report_date = f"{year}/{month}/{day}"
        
        # 漲跌停顏色 (Limit Up/Down)
        limit_up_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        limit_up_font = Font(name='微軟正黑體', size=11, bold=True, color="FFFFFF")
        limit_down_fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        limit_down_font = Font(name='微軟正黑體', size=11, bold=True, color="FFFFFF")
        
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        dark_red_fill = PatternFill(start_color="FF8080", end_color="FF8080", fill_type="solid")
        light_green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        dark_green_fill = PatternFill(start_color="80FF80", end_color="80FF80", fill_type="solid")
        
        date_font = Font(name='微軟正黑體', size=12, bold=True, color="000000")
        header_font = Font(name='微軟正黑體', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        sub_header_font = Font(name='微軟正黑體', size=11, bold=True)
        base_font = Font(name='微軟正黑體', size=11)
        
        # 漲紅跌綠字體
        red_text_font = Font(name='微軟正黑體', size=11, color="FF0000")
        green_text_font = Font(name='微軟正黑體', size=11, color="008800")
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        for market_key, sheet_name in [('TWSE', '上市'), ('TPEX', '上櫃')]:
            ws = wb.create_sheet(title=sheet_name)
            
            # 第一列: 日期
            ws.append([f"{report_date}"])
            ws.cell(row=1, column=1).font = date_font
            ws.cell(row=1, column=1).alignment = left_align
            
            # 篩選出該市場的資料
            market_data = [d for d in all_data if d['market'] == market_key]
            
            # 依買賣超金額排序 (由大到小 / 由深到淺即負數由小到大)
            fb = sorted([d for d in market_data if d['foreign_shares'] > 0], key=lambda x: x['foreign_val'], reverse=True)
            fs = sorted([d for d in market_data if d['foreign_shares'] < 0], key=lambda x: x['foreign_val'])
            ib = sorted([d for d in market_data if d['it_shares'] > 0], key=lambda x: x['it_val'], reverse=True)
            isell = sorted([d for d in market_data if d['it_shares'] < 0], key=lambda x: x['it_val'])
            
            max_rows = max(len(fb), len(fs), len(ib), len(isell))
            
            # 第二列: 大標題 (每區塊 7 欄 + 1 欄位間隔)
            row2 = [
                "外資買超", "", "", "", "", "", "", "",
                "外資賣超", "", "", "", "", "", "", "",
                "投信買超", "", "", "", "", "", "", "",
                "投信賣超", "", "", "", "", "", ""
            ]
            ws.append(row2)
            
            # 第三列: 子標題
            sub_headers = [
                "代號", "名稱", "收盤價", "收盤%", "均價", "張數", "估價(百萬)", "",
                "代號", "名稱", "收盤價", "收盤%", "均價", "張數", "估價(百萬)", "",
                "代號", "名稱", "收盤價", "收盤%", "均價", "張數", "估價(百萬)", "",
                "代號", "名稱", "收盤價", "收盤%", "均價", "張數", "估價(百萬)"
            ]
            ws.append(sub_headers)
            
            # 合併第二列儲存格
            ws.merge_cells("A2:G2")
            ws.merge_cells("I2:O2")
            ws.merge_cells("Q2:W2")
            ws.merge_cells("Y2:AE2")
            
            # 設定前三列樣式
            for cell in ws[2]:
                cell.font = header_font
                cell.alignment = center_align
                if cell.value:
                    cell.fill = header_fill
            for cell in ws[3]:
                cell.font = sub_header_font
                cell.alignment = center_align
                if cell.value:
                    cell.fill = sub_header_fill
            
            # 計算每檔股票的狀態 (同向或反向)
            stock_state = {}
            for d in market_data:
                f_shares, i_shares = d['foreign_shares'], d['it_shares']
                if (f_shares > 0 and i_shares > 0) or (f_shares < 0 and i_shares < 0):
                    if abs(f_shares) > abs(i_shares):
                        stock_state[d['code']] = (dark_red_fill, light_red_fill)
                    else:
                        stock_state[d['code']] = (light_red_fill, dark_red_fill)
                elif (f_shares > 0 and i_shares < 0) or (f_shares < 0 and i_shares > 0):
                    if abs(f_shares) > abs(i_shares):
                        stock_state[d['code']] = (dark_green_fill, light_green_fill)
                    else:
                        stock_state[d['code']] = (light_green_fill, dark_green_fill)
                else:
                    stock_state[d['code']] = (None, None)

            # 輔助函式：取得股票資料與對應的顏色
            def get_stock_data(lst, idx, val_key, lots_key, is_foreign):
                if idx < len(lst):
                    st = lst[idx]
                    fills = stock_state.get(st['code'], (None, None))
                    fill = fills[0] if is_foreign else fills[1]
                    code_val = int(st['code']) if st['code'].isdigit() else st['code']
                    return [code_val, st['name'], st['price'], st['change_pct'], st['vwap'], st[lots_key], st[val_key] / 1000000], fill
                return ["", "", "", "", "", "", ""], None

            # 寫入各分類的排名資料
            for row_i in range(max_rows):
                fb_data, fb_fill = get_stock_data(fb, row_i, 'foreign_val', 'foreign_lots', True)
                fs_data, fs_fill = get_stock_data(fs, row_i, 'foreign_val', 'foreign_lots', True)
                ib_data, ib_fill = get_stock_data(ib, row_i, 'it_val', 'it_lots', False)
                is_data, is_fill = get_stock_data(isell, row_i, 'it_val', 'it_lots', False)
                
                row_idx = row_i + 4 # 標題佔 3 列
                
                col_settings = [
                    (1, fb_data[0], None, None), (2, fb_data[1], fb_fill, None), (3, fb_data[2], None, '#,##0.00'), (4, fb_data[3], None, '0.00"%"'), (5, fb_data[4], None, '#,##0.0'), (6, fb_data[5], None, '#,##0'), (7, fb_data[6], None, '#,##0.00'),
                    (9, fs_data[0], None, None), (10, fs_data[1], fs_fill, None), (11, fs_data[2], None, '#,##0.00'), (12, fs_data[3], None, '0.00"%"'), (13, fs_data[4], None, '#,##0.0'), (14, fs_data[5], None, '#,##0'), (15, fs_data[6], None, '#,##0.00'),
                    (17, ib_data[0], None, None), (18, ib_data[1], ib_fill, None), (19, ib_data[2], None, '#,##0.00'), (20, ib_data[3], None, '0.00"%"'), (21, ib_data[4], None, '#,##0.0'), (22, ib_data[5], None, '#,##0'), (23, ib_data[6], None, '#,##0.00'),
                    (25, is_data[0], None, None), (26, is_data[1], is_fill, None), (27, is_data[2], None, '#,##0.00'), (28, is_data[3], None, '0.00"%"'), (29, is_data[4], None, '#,##0.0'), (30, is_data[5], None, '#,##0'), (31, is_data[6], None, '#,##0.00')
                ]
                
                for col, val, fill, num_fmt in col_settings:
                    if val != "":
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.font = base_font
                        cell.alignment = right_align if isinstance(val, (int, float)) else center_align
                        
                        # 名稱顏色
                        if fill and col in [2, 10, 18, 26]:
                            cell.fill = fill
                            
                        # 漲跌百分比顏色與漲跌停標示
                        if col in [4, 12, 20, 28] and isinstance(val, (int, float)):
                            if val >= 9.5:
                                cell.fill = limit_up_fill
                                cell.font = limit_up_font
                            elif val <= -9.5:
                                cell.fill = limit_down_fill
                                cell.font = limit_down_font
                            elif val > 0:
                                cell.font = red_text_font
                            elif val < 0:
                                cell.font = green_text_font
                        
                        if num_fmt and isinstance(val, (int, float)):
                            cell.number_format = num_fmt

            # 調整欄位寬度
            for c in ['A', 'I', 'Q', 'Y']: # 代號
                ws.column_dimensions[c].width = 8
            for c in ['B', 'J', 'R', 'Z']: # 名稱
                ws.column_dimensions[c].width = 12
            for c in ['C', 'K', 'S', 'AA']: # 收盤價
                ws.column_dimensions[c].width = 9
            for c in ['D', 'L', 'T', 'AB']: # 收盤%
                ws.column_dimensions[c].width = 8
            for c in ['E', 'M', 'U', 'AC']: # 均價
                ws.column_dimensions[c].width = 10
            for c in ['F', 'N', 'V', 'AD']: # 張數
                ws.column_dimensions[c].width = 10
            for c in ['G', 'O', 'W', 'AE']: # 估價
                ws.column_dimensions[c].width = 12
            for c in ['H', 'P', 'X']: # 間隔
                ws.column_dimensions[c].width = 2

        # 新增「統計」分頁 - 土洋買賣超合計
        ws_stat = wb.create_sheet(title="統計")
        ws_stat.append([f"{report_date}"])
        ws_stat.cell(row=1, column=1).font = date_font
        
        # 準備資料
        twse_data = [d for d in all_data if d['market'] == 'TWSE']
        tpex_data = [d for d in all_data if d['market'] == 'TPEX']
        
        # 依照 total_inst_val 排序
        twse_buy = sorted([d for d in twse_data if d['total_inst_val'] > 0], key=lambda x: x['total_inst_val'], reverse=True)
        twse_sell = sorted([d for d in twse_data if d['total_inst_val'] < 0], key=lambda x: x['total_inst_val'])
        tpex_buy = sorted([d for d in tpex_data if d['total_inst_val'] > 0], key=lambda x: x['total_inst_val'], reverse=True)
        tpex_sell = sorted([d for d in tpex_data if d['total_inst_val'] < 0], key=lambda x: x['total_inst_val'])
        
        # 第二列標題
        row2 = ["上市法人合計買超", "", "", "", "", "", "", "", "上市法人合計賣超", "", "", "", "", "", "", "", "上櫃法人合計買超", "", "", "", "", "", "", "", "上櫃法人合計賣超"]
        ws_stat.append(row2)
        for cell in ws_stat[2]:
            cell.font = header_font
            cell.alignment = center_align
            if cell.value: cell.fill = header_fill
        
        # 第三列子標題
        sub_headers = ["代號", "名稱", "收盤價", "漲跌%", "均價", "張數", "估價(M)", ""] * 3 + ["代號", "名稱", "收盤價", "漲跌%", "均價", "張數", "估價(M)"]
        ws_stat.append(sub_headers)
        for cell in ws_stat[3]:
            cell.font = sub_header_font
            cell.alignment = center_align
            if cell.value: cell.fill = sub_header_fill

        # 合併標題
        ws_stat.merge_cells("A2:G2")
        ws_stat.merge_cells("I2:O2")
        ws_stat.merge_cells("Q2:W2")
        ws_stat.merge_cells("Y2:AE2")

        max_rows_stat = max(len(twse_buy), len(twse_sell), len(tpex_buy), len(tpex_sell))
        
        def get_stat_data(lst, idx):
            if idx < len(lst):
                st = lst[idx]
                code_val = int(st['code']) if st['code'].isdigit() else st['code']
                # 合計張數 = foreign_lots + it_lots
                total_lots = st['foreign_lots'] + st['it_lots']
                return [code_val, st['name'], st['price'], st['change_pct'], st['vwap'], total_lots, st['total_inst_val'] / 1000000]
            return ["", "", "", "", "", "", ""]

        for row_i in range(max_rows_stat):
            r_data = []
            r_data.extend(get_stat_data(twse_buy, row_i))
            r_data.append("")
            r_data.extend(get_stat_data(twse_sell, row_i))
            r_data.append("")
            r_data.extend(get_stat_data(tpex_buy, row_i))
            r_data.append("")
            r_data.extend(get_stat_data(tpex_sell, row_i))
            
            ws_stat.append(r_data)
            row_idx = row_i + 4
            for col_idx, val in enumerate(r_data, 1):
                if val != "":
                    cell = ws_stat.cell(row=row_idx, column=col_idx)
                    cell.font = base_font
                    cell.alignment = right_align if isinstance(val, (int, float)) else center_align
                    if col_idx in [4, 12, 20, 28] and isinstance(val, (int, float)):
                        if val > 0: cell.font = red_text_font
                        elif val < 0: cell.font = green_text_font
                    if isinstance(val, (int, float)):
                        if col_idx in [4, 12, 20, 28]: cell.number_format = '0.00"%"'
                        elif col_idx in [3, 5, 7, 11, 13, 15, 19, 21, 23, 27, 29, 31]: cell.number_format = '#,##0.0'

        # 調整欄寬 (比照主分頁)
        stat_cols = ['A', 'I', 'Q', 'Y', 'B', 'J', 'R', 'Z', 'C', 'K', 'S', 'AA', 'D', 'L', 'T', 'AB', 'E', 'M', 'U', 'AC', 'F', 'N', 'V', 'AD', 'G', 'O', 'W', 'AE']
        for c in stat_cols:
            if c in ['A', 'I', 'Q', 'Y']: ws_stat.column_dimensions[c].width = 8
            elif c in ['B', 'J', 'R', 'Z']: ws_stat.column_dimensions[c].width = 12
            else: ws_stat.column_dimensions[c].width = 10

        # filename is already set above
        wb.save(filename)
        print(f"\n已成功輸出多欄位變色 Excel 報表: {filename}")
        return True
    except Exception as e:
        print(f"\n輸出報表時發生錯誤: {e}")
        return False

if __name__ == '__main__':
    import sys
    import argparse
    
    parser = argparse.ArgumentParser(description='Taiwan Stock Market Institutional Investor Analysis')
    parser.add_argument('date', nargs='?', help='Target date (YYYYMMDD)')
    parser.add_argument('--period', choices=['day', 'week'], default='day', help='Analysis period')
    args = parser.parse_args()
    
    input_date = args.date
    period = args.period
    
    # 無論是有輸入日期還是自動觸發，如果發現當天沒開盤，都應該回溯尋找
    success = False
    
    # 決定起始日期
    if input_date:
        # 使用者指定的日期 (格式 YYYYMMDD)
        start_date_obj = datetime.strptime(input_date, '%Y%m%d')
        print(f"User requested {period} analysis starting from: {input_date}")
    else:
        # 自動模式，從今天開始找
        start_date_obj = datetime.now()
        print(f"Automatic {period} trigger starting from today...")

    # 智慧回溯循環 (最多往回找 10 天交易日)
    for i in range(10):
        current_date_str = (start_date_obj - timedelta(days=i)).strftime('%Y%m%d')
        print(f"--- [快速預檢] 測試日期: {current_date_str} (Day {i+1}) ---")
        
        if validate_trading_day(current_date_str):
            print(f"[OK] 成功命中有效交易日: {current_date_str}！ 準備開始執行重型分析任務...")
            try:
                if analyze(current_date_str, period=period):
                    success = True
                    break
            except Exception as e:
                print(f"[ERROR] 執行分析時發生非預期錯誤: {e}")
                traceback.print_exc()
                break
        else:
            print(f"[WARN] 日期 {current_date_str} 休市中，自動跳過...")
            continue
    
    if not success:
        print("[CRITICAL] 任務失敗：在最近的 10 天內找不到任何開盤紀錄，請檢查證交所連線或網站狀態。")
        sys.exit(1)
